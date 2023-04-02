from warehouse_entry import WarehouseEntry
from openpyxl import load_workbook, Workbook

def export_entry_workbook(entry : WarehouseEntry) -> Workbook:
    wb = load_workbook('./template.xlsx')

    ws = wb.active
    ws['G2'] = entry.supplier_name
    ws['C3'] = entry.invoice_code
    ws['G3'] = entry.invoice_number
    ws['C4'] = entry.purchase_date
    ws['G4'] = entry.entry_date
    ws['C18'] = entry.amount_with_tax_zh
    ws['H18'] = entry.amount_with_tax
    row_num = 6
    col_num = 3
    for item in entry.items:
        for i in range(7):
            ws.cell(row=row_num, column=col_num+i, value=item.line_content[i])
        row_num += 1

    # wb.save('entry.xlsx')
    return wb
